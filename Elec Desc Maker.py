from openpyxl import load_workbook

wb = load_workbook("exampleSheet.xlsx")
mySheet = wb.worksheets[0]



for i in range(1, mySheet.max_row + 1, 1):
    # Variable holds the assetID value = "FS10-2-B1 N-L-PP3-S1F/S2F"
    assetID = mySheet.cell(i, 3).value

    # This list will hold the assetID split by the space in the middle = "FS10-2-B1", "N-L-PP3-S1F/S2F"
    IDParts = assetID.split(' ')



    # IDParts is a list that should have 2 elements, but not all rows will be complete
    if len(IDParts) == 2:

        # This will split the 2nd half of the assetID by the dashes '-' = "N", "L", "PP3", "S1F/S2F"
        elecParts = IDParts[1].split('-')

        # Correct rows will have 4 elements in elecParts
        if len(elecParts) == 4:

            # Power type
            powerType = elecParts[0].upper()

            match powerType:
                case 'N':
                    powerType = "NORMAL"
                
                case 'L':
                    powerType = "LOW"

                case 'S':
                    powerType = "LIFE SAFETY"
                    
                case 'E':
                    powerType = "EMERGENCY"

                case 'R':
                    powerType = "LEGALLY REQUIRED"

                case 'ERS':
                    powerType = "EMERGENCY, LEGALLY REQUIRED, LIFE SAFEY"

                case 'U':
                    powerType = "UPS"

                # If something is wrong, this will print the switch statement where it went wrong
                case _:
                    print(f"Row {i}: ElecParts[0]")

            

            # Voltages
            voltage = elecParts[1].upper()
            match voltage:
                case "H":
                    voltage = "480VAC"

                case "L":
                    voltage = "277VAC"

                case "M":
                    voltage = "ABOVE 480VAC"

                case "B":
                    voltage = "120/240VAC"

                case "O":
                    voltage = "24VDC AND BELOW"

                # If something is wrong, this will print the switch statement where it went wrong
                case _:
                    print(f"Row {i}: ElecParts[1]")



            # Equipment Type
            EType = elecParts[2].upper()

            # Position and class all come from the equipment type below, so we set those at the same time
            pos = ""
            assetClass = ""

            # if EType is 3 long, that means its something like PP1, SB3, etc.
            # FPC is the only EType that would be 4
            if len(EType) == 3:
                match EType[:2]:
                    case "PP":
                        EType = "POWER PANEL"
                        pos = "FC.E.DIST.PP"
                        assetClass = "EDIST.PP"
                        
                    case "TR":
                        EType = "TRANSFORMER"
                        pos = "FC.E.DIST.XFMR"
                        assetClass = "XFMR"

                    case "AT":
                        EType = "AUTOMATIC TRANSFER SWITCH"
                        pos = "FC.E.DIST.ATS"
                        assetClass = "EDIST.AT"

                    case "DS":
                        EType = "DISCONNECT.NONFUSED"
                        pos = "FC.E.DIST.DIS"
                        assetClass = "DISC"

                    case "SB":
                        EType = "SWITCHBOARD"
                        pos = "FC.E.DIST.SB"
                        assetClass = "EDIST.SB"

                    case "LC":
                        EType = "LIGHTING CONTROL"
                        pos = "FC.E.DIST.LIGHTING"
                        assetClass = "EDIST.LP"
                        
                    case "SG":
                        EType = "SWITCHGEAR"
                        pos = "FC.E.DIST.SG"
                        assetClass = "EDIST.SG"
                        
                    case "DF":
                        EType = "DISCONNECT.FUSED"
                        pos = "FC.E.DIST.DIS"
                        assetClass = "DISC"
                        
                    case "MS":
                        EType = "MOTOR STARTER"
                        pos = "FC.E.DIST.MS"
                        assetClass = "EDIST.MS"
                        
                    case "MC":
                        EType = "MOTOR CONTROL CENTER"
                        pos = "FC.E.DIST.MCC"
                        assetClass = "EDISTMCC"
                        
                    case "UP":
                        EType = "UPS UNIT"
                        pos = "FC.E.DIST.UPS"
                        assetClass = "UPS"

                    # If something is wrong, this will print the switch statement where it went wrong
                    case _:
                        print(f"Row {i}: ElecParts[2]")

            # If first 3 chars of EType is FPC, do the following
            # This is the only entry that is 3 chars
            elif EType[:3] == "FPC":
                EType = "FIRE PUMP CONTROLLER"
                pos = "FC.FIRE.SUP.WET.FIRE.PMP.FCU"
                assetClass = "FCU"



            # The last part is the service sources
            serviceSource = elecParts[3]

            # Combine the parts to create the full description
            fullDesc = f"{EType}.{powerType}.{voltage}.{serviceSource}"

            # Set the values to be the new description
            mySheet.cell(i, 5).value = fullDesc

            # Set the position to the correct value
            mySheet.cell(i, 2).value = pos

            # Set the class
            mySheet.cell(i, 11).value = assetClass



# Save the file at the end
wb.save("exampleSheet.xlsx")